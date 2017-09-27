import csv
import random
import matplotlib.pyplot as plt
import numpy
import collections
import os
import monitor_space
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Color,PatternFill,Border,Side



def calculateGrowthWithTests(days):
    """
    This function will calculate db growth with the fake test files I made previously
    """
    machineDict={}
    for x in range(days):
        fileName=("dbtest"+str(x)+".csv")
        #Set up dictionary with values for all days
        with open (fileName) as f:
            for line in f:
                line=line.split(',')
                line[0]=line[0].replace('\'',' ').strip().replace('(','').strip()
                if(len(line)>1):
                    line[1]=round(float(line[1].replace(')',' ').strip()),2)
                try:
                    if(line[0] not in machineDict):
                        machineDict[line[0]]=[line[1]]
                    else:
                        machineDict[line[0]].append(line[1])
                except:
                    pass
    machineDictPercentChange={}
    #Make a dictionary with percent change from each day
    for key in machineDict:
        machineDictPercentChange[key]=[]
        values=machineDict[key]
        for x in range (len(values)-1):
            if(values[x]!=0):
                change=(values[x]-values[x+1])/values[x]*100
                change=round(change,2)
                machineDictPercentChange[key].append(change)
            else:
                machineDictPercentChange[key].append(0.0)
    machineDictGrowthEstimate={}
    for key in machineDictPercentChange:
        #Calculate growth guesstimate
        try:
            """
            #((most recent period growth/starting period growth)**(1-number of periods))-1
            X=(machineDictPercentChange[key][-1])/(machineDictPercentChange[key][0])
            Y=days-1
            growth=((X)**(1/Y))-1
            machineDictGrowthEstimate[key]=round(growth,2)
            """
            averageGrowth=sum(machineDictPercentChange[key])/len(machineDictPercentChange[key])
            machineDictGrowthEstimate[key]=round(averageGrowth,2)
        except:
            raise
    for key in machineDictGrowthEstimate:
        print(key,machineDictGrowthEstimate[key])
    plotEntireDB(machineDictGrowthEstimate,machineDict,days)

def calculateGrowthWithFacts(files):
    """
    This function will calculate db growth legitimate files that have factual values from the DB
    """
    machineDict={}
    for line in files:
        fileName=(line)
        #Set up dictionary with values for all days
        with open (fileName) as f:
            for line in f:
                line=line.split(',')
                line[0]=line[0].replace('\'',' ').strip().replace('(','').strip()
                if(len(line)>1):
                    line[1]=round(float(line[1].replace(')',' ').strip()),2)
                try:
                    if(line[0] not in machineDict):
                        machineDict[line[0]]=[line[1]]
                    else:
                        machineDict[line[0]].append(line[1])
                except:
                    pass
    machineDictPercentChange={}
    #Make a dictionary with percent change from each day
    for key in machineDict:
        machineDictPercentChange[key]=[]
        values=machineDict[key]
        for x in range (len(values)-1):
            if(values[x]!=0):
                change=(values[x]-values[x+1])/values[x]*100
                change=round(change,2)
                machineDictPercentChange[key].append(change)
            else:
                machineDictPercentChange[key].append(0.0)
    machineDictGrowthEstimate={}
    for key in machineDictPercentChange:
        #Calculate growth guesstimate
        try:
            """
            #((most recent period growth/starting period growth)**(1-number of periods))-1
            X=(machineDictPercentChange[key][-1])/(machineDictPercentChange[key][0])
            Y=days-1
            growth=((X)**(1/Y))-1
            machineDictGrowthEstimate[key]=round(growth,2)
            """
            averageGrowth=sum(machineDictPercentChange[key])/len(machineDictPercentChange[key])
            machineDictGrowthEstimate[key]=round(averageGrowth,2)
        except:
            raise
    for key in machineDictGrowthEstimate:
        print(key,machineDictGrowthEstimate[key])
    plotEntireDB(machineDictGrowthEstimate,machineDict,len(files))

def plotEntireDB(growthEstimate,machineDict,days):
    """
    This function will plot the entire database on a graph for however many days worth of files were
    given. Afterwards, the function will get the average growth of loss of the database and ask the User
    for however many days they would like to predict in the future. The function will then plot what
    the expected outcome will be.
    """
    #machineDict=sorted(machineDict, key=machineDict.get, reverse=True)
    dailyTotal=[]
    for x in range(0,days):
        totalUsage=0
        for item in machineDict:
            try:
                totalUsage+=(float(machineDict[item][x]))
            except:
                pass
        dailyTotal.append(totalUsage)
    try:
        plt.scatter(list(range(1,days+1)),dailyTotal,label="Database")
        plt.plot(list(range(1,days+1)),dailyTotal)
        plt.xlabel("Days")
        plt.ylabel("Mbs")
        plt.legend()
        plt.savefig('overall_database_usage.png', bbox_inches='tight')
        #plt.show()
    except UserWarning:
        pass
    machineDictPercentChange=[]
    machineDictMBChange=[]
    for x in range(len(dailyTotal)-1):
        changePercent=(dailyTotal[x]-dailyTotal[x+1])/dailyTotal[x]*100
        changePercent=round(changePercent,2)
        machineDictPercentChange.append(changePercent)
    for x in range(len(dailyTotal)-1):
        mbChange=dailyTotal[x]-dailyTotal[x+1]
        machineDictMBChange.append(mbChange)
    averageMBchange=numpy.mean(machineDictMBChange)
    averageGrowth=numpy.mean(machineDictPercentChange)
    predictionTotal=dailyTotal
    daysPrediction=int(input("How many days would you like to predict growth for? "))
    for x in range(int(daysPrediction)):
        predictionTotal.append(predictionTotal[-1]-(averageMBchange))
    plt.scatter(list(range(1,days+1+int(daysPrediction))),predictionTotal,label="Database")
    plt.plot(list(range(1,days+1+int(daysPrediction))),predictionTotal)
    plt.xlabel("Days")
    plt.ylabel("Mbs")
    plt.legend()
    plt.savefig('overall_database_prediction.png', bbox_inches='tight')
    #plt.show()
    capacity=capacityEstimate(machineDict,averageMBchange)
    pivotTable(machineDict,daysPrediction,days,capacity)

def capacityEstimate(machineDict,averageMBChange):
    """
    This function figures out based on the current trend of growth, when the database will reach certain
    capacity limits
    """
    total=0
    averageMBChange=averageMBChange*-1
    for item in machineDict:
        total+=(float(machineDict[item][-1]))
    print(total)
    dayEstimates=[total]
    daysAhead=0
    while(total<2000000):
        total+=averageMBChange
        daysAhead+=1
    dayEstimates.append(daysAhead)
    while(total<2500000):
        total+=averageMBChange
        daysAhead+=1
    dayEstimates.append(daysAhead)
    while(total<3000000):
        total+=averageMBChange
        daysAhead+=1
    dayEstimates.append(daysAhead)
    return dayEstimates

def averageMBgrowth(machine):
    """
    This function calulates the average growth of a group of machines
    """
    averageMachineGrowth={}
    for item in machine:
        mbChanges=[]
        for x in range(len(machine[item])-1):
            mbChange=machine[item][x]-machine[item][x+1]
            mbChanges.append(mbChange)
        averageMachineGrowth[item]=numpy.mean(mbChanges)
    return averageMachineGrowth

def pivotTable(machineDict,daysPrediction,days,dayEstimates):
    """
    This function creates the outcome for the entire program. It creates a pivot table with 4 sheets
    1. The machines with the highest MB use. It will show the average use over X amount of days and then give a prediction of what the machine will look like for X amount of days afterwards
    2. A graph of the overall database usage
    3. A graph of the expected database usage in the future
    4. Excel sheet of the current capacity and expected days until certain other capacities.
    """
    #sorted(machineDict.items(), key=lambda x:x[1])
    averageUsage={}
    for item in machineDict:
        averageUsage[item]=numpy.mean(machineDict[item])
    sortedValues=sorted(averageUsage.values())
    sortedAverageUsage={}
    for i in sortedValues[-19:]:
        for item in averageUsage:
            if (averageUsage[item]==i):
                sortedAverageUsage[item]=i
    averageGrowthDict=averageMBgrowth(machineDict)
    sortedAverageGrowth={}
    for item in sortedAverageUsage:
        if (item in averageGrowthDict):
            sortedAverageGrowth[item]=averageGrowthDict[item]
    wb=Workbook()
    ws=wb.create_sheet("Highest Users",0)
    y="Average mb growth over "+str(days)+" days"
    cols=["Name", y]
    for i in range(daysPrediction+1):
        cols.append("Day "+str(days+i))
    ws.append(cols)
    for item in sortedAverageGrowth:
        machineEstimate=[item,float(sortedAverageGrowth[item])*-1,machineDict[item][-1]]
        for i in range(daysPrediction):
            machineEstimate.append(machineEstimate[-1]-sortedAverageGrowth[item])
        ws.append(machineEstimate)
    adjustColWidth(ws)
    ws.append([])
    cols=['Current Capacity','Days Expected to reach: Normal-2TB','Critical-2.5TB','Major>=3TB']
    ws.append(cols)
    ws.append(dayEstimates)
    """
    ws=wb.create_sheet("Overall Database Usage",1)
    ws=wb.worksheets[1]
    """
    img = Image('overall_database_usage.png')
    ws.add_image(img, 'A25')
    """
    ws=wb.create_sheet("Overall Database Prediction",2)
    ws=wb.worksheets[2]
    """
    img = Image('overall_database_prediction.png')
    ws.add_image(img, 'A48')
    """
    ws=wb.create_sheet("Expected Capacity Measurements",3)
    ws=wb.worksheets[3]
    """

    adjustColWidth(ws)
    addColor(ws)
    wb.save("DB Pivot Table.xlsx")

def addColor(sheet):
    #Set up colors
    blueFill = PatternFill(start_color='00FFFF',
                   end_color='00FFFF',
                   fill_type='solid')
    greenFill = PatternFill(start_color='00FF00',
                   end_color='00FFFF',
                   fill_type='solid')
    yellowFill = PatternFill(start_color='FFFF00',
                   end_color='00FFFF',
                   fill_type='solid')
    redFill = PatternFill(start_color='FF0000',
                   end_color='00FFFF',
                   fill_type='solid')
    sheet['A1'].fill=blueFill
    sheet['B1'].fill=yellowFill
    for col in sheet.iter_cols(min_row=1, min_col=3,max_col=len(sheet['1:']), max_row=1):
        for cell in col:
            cell.fill=greenFill
    sheet['A22'].fill=blueFill
    sheet['B22'].fill=greenFill
    sheet['C22'].fill=yellowFill
    sheet['D22'].fill=redFill

def plotPoints(growthEstimate,machineDict,days):
    """
    Show a graph of every machine and it's usage, with 10 on a page at a time
    """
    count=0
    for item in sorted(machineDict, key=machineDict.get, reverse=True):
        plt.scatter(list(range(1,days+1)),machineDict[item],label=item)
        plt.plot(range(1,days+1),machineDict[item])
        plt.xlabel("Days")
        plt.ylabel("Mbs")
        plt.title(item)
        plt.legend()
        if(count==10):
            plt.show()
            count=0
        count+=1

def adjustColWidth(sheet):
    """
    This function adjust the width of a column for a better looking excel sheet
    """
    thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
    for col in sheet.columns:
        max_length = 6
        column = col[0].column # Get the column name
        for cell in col:
            cell.border=thin_border
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width
    return sheet

def getFiles():
    """
    Search the current working directory for all files that start with 'dbsizes'. All files with that
    contain information about the databases.
    """
    directory=os.listdir()
    files=[]
    for item in directory:
        if (item.startswith("dbsizes")):
            files.append(item)
    return files

def getOldestFile(files):
    oldestTime=0.0
    print(files)
    for item in files:
        if(os.path.getmtime(item)>oldestTime):
            print(item,os.path.getmtime(item),oldestTime)
            oldestTime=os.path.getmtime(item)
            input()
            oldestFile=item

    print(item)
    input()
    return item
def updateFiles():
    #Check if folders for files already exist, if they don't create them
    if not os.path.exists("current_DB_files"):
        os.makedirs("current_DB_files")
    if not os.path.exists("archived_DB_files"):
        os.makedirs("archived_DB_files")
    #get current db files
    os.chdir("current_DB_files")
    currentFiles=getFiles()
    getOldestFile(currentFiles)
    #os.rename(getOldestFile(currentFiles), "../archived_DB_files/"+getOldestFile(currentFiles))






    """
    os.chdir("../archived_DB_files")
    archivedFiles=getFiles()
    os.chdir("..")
    monitor_space.main()
    """



def main():
    #days=int(input("How many days of data should be made? "))
    #createTimeFrame(days)

    #return list from getFiles
    updateFiles()
    #calculateGrowthWithFacts(getFiles())





if __name__ =='__main__':
    main()
